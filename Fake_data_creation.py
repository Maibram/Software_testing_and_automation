from faker import Faker
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

fake_data = Faker()

for  i in range(1, 11):
    for j in range(1, 4):
        ws.cell(row=i, column=j).value = fake_data.name()
        ws.cell(row=i, column=j+1).value = fake_data.email()
        ws.cell(row=i, column=j+2).value = fake_data.phone_number()
        ws.cell(row=i, column=j+3).value = fake_data.address()